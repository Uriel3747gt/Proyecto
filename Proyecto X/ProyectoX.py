import tkinter as tk
from tkinter import messagebox, simpledialog
import sqlite3
import openpyxl
import webbrowser


class Inventario:
    def __init__(self, db_path='Inventario.db'):
        self.conn = sqlite3.connect(db_path)
        self.cursor = self.conn.cursor()
        self._crear_tabla_si_no_existe()
        self.clientes = Cliente()

    def _crear_tabla_si_no_existe(self):
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS productos (
                id INTEGER PRIMARY KEY,
                nombre TEXT,
                cantidad INTEGER
            )
        ''')
        self.conn.commit()

    def agregar_producto(self, nombre, cantidad):
        self.cursor.execute("SELECT * FROM productos WHERE nombre=?", (nombre,))
        existing_product = self.cursor.fetchone()

        if existing_product:
            nueva_cantidad = existing_product[2] + cantidad
            self.cursor.execute("UPDATE productos SET cantidad=? WHERE id=?", (nueva_cantidad, existing_product[0]))
        else:
            self.cursor.execute("INSERT INTO productos (nombre, cantidad) VALUES (?, ?)", (nombre, cantidad))

        self.conn.commit()

    def eliminar_producto(self, nombre):
        self.cursor.execute("DELETE FROM productos WHERE nombre=?", (nombre,))
        self.conn.commit()

    def actualizar_producto(self, nombre, nueva_cantidad):
        self.cursor.execute("SELECT * FROM productos WHERE nombre=?", (nombre,))
        existing_product = self.cursor.fetchone()

        if existing_product:
            self.cursor.execute("UPDATE productos SET cantidad=? WHERE id=?", (nueva_cantidad, existing_product[0]))
            self.conn.commit()
        else:
            messagebox.showwarning("Error", "Producto no encontrado")

    def mostrar_inventario(self):
        self.cursor.execute("SELECT * FROM productos")
        inventario = self.cursor.fetchall()

        inventario_texto = "Inventario:\n"
        for producto in inventario:
            inventario_texto += f"{producto[1]}: {producto[2]}\n"

        messagebox.showinfo("Inventario", inventario_texto)

    def exportar_a_excel(self, nombre_archivo):
        self.cursor.execute("SELECT * FROM productos")
        inventario = self.cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = 'Producto'
        ws['B1'] = 'Cantidad'

        for row, producto in enumerate(inventario, start=2):
            ws.cell(row=row, column=1, value=producto[1])
            ws.cell(row=row, column=2, value=producto[2])

        wb.save(nombre_archivo)
        messagebox.showinfo("Exportado", f"Datos exportados a {nombre_archivo}")

    def enviar_correo(self, destinatario, asunto, cuerpo):
        # Abre el cliente de correo predeterminado con un mensaje predefinido
        mailto_link = f"mailto:{destinatario}?subject={asunto}&body={cuerpo}"

        try:
            webbrowser.open(mailto_link)
            messagebox.showinfo("Correo Enviado", f"Se abrió el cliente de correo para enviar a {destinatario}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el cliente de correo. Error: {e}")


class Cliente:
    def __init__(self):
        self.clientes = {}

    def crear_cliente(self, nombre, email):
        self.clientes[nombre] = {'email': email, 'compras': 0}

    def agregar_cliente(self, nombre, compras):
        if nombre in self.clientes:
            self.clientes[nombre]['compras'] += compras
        else:
            messagebox.showwarning("Error", "Cliente no encontrado")

    def actualizar_cliente(self, nombre, email):
        if nombre in self.clientes:
            self.clientes[nombre]['email'] = email
        else:
            messagebox.showwarning("Error", "Cliente no encontrado")

    def eliminar_cliente(self, nombre):
        if nombre in self.clientes:
            del self.clientes[nombre]
        else:
            messagebox.showwarning("Error", "Cliente no encontrado")

    def mostrar_clientes(self):
        clientes_texto = "Clientes:\n"
        for cliente, info in self.clientes.items():
            clientes_texto += f"{cliente} - Email: {info['email']} - Compras: {info['compras']}\n"
        messagebox.showinfo("Clientes", clientes_texto)

    def exportar_a_excel(self, nombre_archivo):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = 'Cliente'
        ws['B1'] = 'Email'
        ws['C1'] = 'Compras'

        row = 2
        for cliente, info in self.clientes.items():
            ws.cell(row=row, column=1, value=cliente)
            ws.cell(row=row, column=2, value=info['email'])
            ws.cell(row=row, column=3, value=info['compras'])
            row += 1

        wb.save(nombre_archivo)
        messagebox.showinfo("Exportado", f"Datos exportados a {nombre_archivo}")


class InterfazInventario:
    def __init__(self, inventario):
        self.inventario = inventario
        self.master = tk.Tk()
        self.master.title("Inventario y Clientes")

        self.frame = tk.Frame(self.master)
        self.frame.pack()

        self.label = tk.Label(self.frame, text="Comandos:")
        self.label.grid(row=0, column=0, columnspan=2)

        self.opciones_menu = tk.StringVar(self.master)
        self.opciones_menu.set("Seleccionar opción")

        self.menu_desplegable = tk.OptionMenu(
            self.frame, self.opciones_menu, "Seleccionar opción", "Agregar Producto",
            "Eliminar Producto", "Actualizar Producto", "Mostrar Inventario",
            "Crear Cliente", "Agregar Compras a Cliente", "Actualizar Cliente",
            "Eliminar Cliente", "Mostrar Clientes", "Exportar Inventario a Excel",
            "Exportar Clientes a Excel", "Enviar Correo"
        )
        self.menu_desplegable.grid(row=1, column=0, columnspan=2)

        self.btn_ejecutar = tk.Button(self.frame, text="Ejecutar", command=self.ejecutar_opcion)
        self.btn_ejecutar.grid(row=2, column=0, columnspan=2)

    def ejecutar_opcion(self):
        opcion = self.opciones_menu.get()

        if opcion == "Seleccionar opción":
            messagebox.showwarning("Error", "Por favor, seleccione una opción")
        elif opcion == "Agregar Producto":
            producto = simpledialog.askstring("Agregar Producto", "Ingresar nuevo producto:")
            cantidad = simpledialog.askinteger("Agregar Cantidad", "Ingrese la cantidad a agregar:")
            self.inventario.agregar_producto(producto, cantidad)
        elif opcion == "Eliminar Producto":
            producto = simpledialog.askstring("Eliminar Producto", "Ingrese el nombre del producto a eliminar:")
            self.inventario.eliminar_producto(producto)
        elif opcion == "Actualizar Producto":
            producto = simpledialog.askstring("Actualizar Producto", "Ingresar nombre del producto a actualizar:")
            cantidad = simpledialog.askinteger("Actualizar Cantidad", "Ingrese la nueva cantidad:")
            self.inventario.actualizar_producto(producto, cantidad)
        elif opcion == "Mostrar Inventario":
            self.inventario.mostrar_inventario()
        elif opcion == "Crear Cliente":
            nombre = simpledialog.askstring("Crear Cliente", "Ingresar nuevo nombre:")
            correo = simpledialog.askstring("Crear Cliente", "Ingrese el nuevo correo electrónico:")
            self.inventario.clientes.crear_cliente(nombre, correo)
        elif opcion == "Agregar Compras a Cliente":
            nombre = simpledialog.askstring("Agregar Compras a Cliente", "Ingrese el nombre del cliente:")
            compras = simpledialog.askinteger("Agregar Compras a Cliente", "Ingrese la cantidad de compras:")
            self.inventario.clientes.agregar_cliente(nombre, compras)
        elif opcion == "Actualizar Cliente":
            nombre = simpledialog.askstring("Actualizar Cliente", "Ingrese el nombre del cliente:")
            correo = simpledialog.askstring("Actualizar Cliente", "Ingrese el nuevo correo electrónico:")
            self.inventario.clientes.actualizar_cliente(nombre, correo)
        elif opcion == "Eliminar Cliente":
            nombre = simpledialog.askstring("Eliminar Cliente", "Ingrese el nombre del cliente a eliminar:")
            self.inventario.clientes.eliminar_cliente(nombre)
        elif opcion == "Mostrar Clientes":
            self.inventario.clientes.mostrar_clientes()
        elif opcion == "Exportar Inventario a Excel":
            archivo = simpledialog.askstring("Exportar Inventario a Excel", "Ingrese el nombre del archivo Excel:")
            self.inventario.exportar_a_excel(archivo)
        elif opcion == "Exportar Clientes a Excel":
            archivo = simpledialog.askstring("Exportar Clientes a Excel", "Ingrese el nombre del archivo Excel:")
            self.inventario.clientes.exportar_a_excel(archivo)
        elif opcion == "Enviar Correo":
            destinatario = simpledialog.askstring("Enviar Correo", "Ingrese la dirección de correo del destinatario:")
            asunto = simpledialog.askstring("Enviar Correo", "Ingrese el asunto del correo:")
            cuerpo = simpledialog.askstring("Enviar Correo", "Ingrese el cuerpo del correo:")
            self.inventario.enviar_correo(destinatario, asunto, cuerpo)
        else:
            messagebox.showwarning("Error", "Opción no válida")

    def run(self):
        self.master.mainloop()


def ejecutar_interfaz():
    inventario = Inventario()
    app = InterfazInventario(inventario)
    app.run()


def ejecutar_consola():
    inventario = Inventario()
    clientes = Cliente()

    while True:
        print("Menú Principal:")
        print("1. Agregar Producto")
        print("2. Eliminar Producto")
        print("3. Actualizar Producto")
        print("4. Mostrar Inventario")
        print("5. Crear Cliente")
        print("6. Agregar Compras a Cliente")
        print("7. Actualizar Cliente")
        print("8. Eliminar Cliente")
        print("9. Mostrar Clientes")
        print("10. Exportar Inventario a Excel")
        print("11. Exportar Clientes a Excel")
        print("12. Enviar Correo")
        print("0. Salir")

        opcion = input("Ingrese el número de la opción deseada: ")

        if opcion == "0":
            print("¡Hasta luego!")
            break
        elif opcion == "1":
            producto = input("Ingresar nuevo producto: ")
            cantidad = int(input("Ingrese la cantidad a agregar: "))
            inventario.agregar_producto(producto, cantidad)
        elif opcion == "2":
            producto = input("Ingrese el nombre del producto a eliminar: ")
            inventario.eliminar_producto(producto)
        elif opcion == "3":
            producto = input("Ingresar nombre del producto a actualizar: ")
            cantidad = int(input("Ingrese la nueva cantidad: "))
            inventario.actualizar_producto(producto, cantidad)
        elif opcion == "4":
            inventario.mostrar_inventario()
        elif opcion == "5":
            nombre = input("Ingresar nuevo nombre: ")
            correo = input("Ingrese el nuevo correo electrónico: ")
            clientes.crear_cliente(nombre, correo)
        elif opcion == "6":
            nombre = input("Ingrese el nombre del cliente: ")
            compras = int(input("Ingrese la cantidad de compras: "))
            clientes.agregar_cliente(nombre, compras)
        elif opcion == "7":
            nombre = input("Ingrese el nombre del cliente: ")
            correo = input("Ingrese el nuevo correo electrónico: ")
            clientes.actualizar_cliente(nombre, correo)
        elif opcion == "8":
            nombre = input("Ingrese el nombre del cliente a eliminar: ")
            clientes.eliminar_cliente(nombre)
        elif opcion == "9":
            clientes.mostrar_clientes()
        elif opcion == "10":
            archivo = input("Ingrese el nombre del archivo Excel: ")
            inventario.exportar_a_excel(archivo)
        elif opcion == "11":
            archivo = input("Ingrese el nombre del archivo Excel: ")
            clientes.exportar_a_excel(archivo)
        elif opcion == "12":
            destinatario = input("Ingrese la dirección de correo del destinatario: ")
            asunto = input("Ingrese el asunto del correo: ")
            cuerpo = input("Ingrese el cuerpo del correo: ")
            inventario.enviar_correo(destinatario, asunto, cuerpo)
        else:
            print("Opción no válida. Por favor, ingrese un número válido.")


def modo_interactivo():
    while True:
        print("Seleccione el modo de ejecución:")
        print("1. Interfaz Gráfica")
        print("2. Consola")
        print("0. Salir")

        opcion = input("Ingrese el número de la opción deseada: ")

        if opcion == "0":
            print("¡Hasta luego!")
            break
        elif opcion == "1":
            ejecutar_interfaz()
        elif opcion == "2":
            ejecutar_consola()
        else:
            print("Opción no válida. Por favor, ingrese un número válido.")


if __name__ == "__main__":
    modo_interactivo()
